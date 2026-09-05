#!/usr/bin/env python3
"""
Export de la base outbound consolidée en fichier .xlsx (ouverture directe dans
Google Sheets / Excel). Lit tous les *-exposants.csv et produit base-outbound.xlsx
avec mise en forme : en-tête figé, filtres, largeurs, onglet de synthèse.

Run: python3 export_sheet.py
Out: base-outbound.xlsx
"""
import csv, glob, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

COLS = ["salon", "edition_annee", "entreprise", "secteur_metier", "site_web", "email",
        "type_email", "email_origine", "email_annuaire", "email_site", "email_site_source",
        "nom_contact", "fonction", "ville", "departement", "telephone",
        "source_url", "date_collecte", "statut_validation", "segment_ICP"]
HEAD_FR = {"salon": "Salon", "edition_annee": "Édition", "entreprise": "Entreprise",
           "secteur_metier": "Secteur / métier", "site_web": "Site web", "email": "Email",
           "type_email": "Type email", "email_origine": "Origine", "email_annuaire": "Email annuaire",
           "email_site": "Email site", "email_site_source": "Source email (URL)",
           "nom_contact": "Contact", "fonction": "Fonction", "ville": "Ville",
           "departement": "Dépt", "telephone": "Téléphone", "source_url": "Source",
           "date_collecte": "Date collecte", "statut_validation": "Statut", "segment_ICP": "ICP"}


def load_rows():
    rows = []
    for path in sorted(glob.glob("*-exposants.csv")):
        for r in csv.DictReader(open(path, encoding="utf-8"), delimiter=";"):
            rows.append({c: r.get(c, "") for c in COLS})
    return rows


def main():
    rows = load_rows()
    wb = Workbook()
    # --- Onglet Synthèse ---
    s = wb.active; s.title = "Synthèse"
    salons = sorted({r["salon"] for r in rows})
    n_mail = sum(1 for r in rows if r["email"])
    n_icp = sum(1 for r in rows if r["segment_ICP"] == "Haute")
    n_ok = sum(1 for r in rows if r["statut_validation"] == "valide")
    s.append(["Base outbound — synthèse"]); s["A1"].font = Font(bold=True, size=14)
    s.append([])
    for k, v in [("Exposants", len(rows)), ("Emails trouvés", n_mail),
                 ("Dont statut 'valide'", n_ok), ("ICP Haute", n_icp),
                 ("Salons", len(salons))]:
        s.append([k, v])
    s.append([])
    s.append(["Détail par salon", "Exposants", "Emails"]); s["A9"].font = Font(bold=True)
    for sa in salons:
        sub = [r for r in rows if r["salon"] == sa]
        s.append([sa, len(sub), sum(1 for r in sub if r["email"])])
    s.column_dimensions["A"].width = 26; s.column_dimensions["B"].width = 14
    s.column_dimensions["C"].width = 12

    # --- Onglet Exposants ---
    ws = wb.create_sheet("Exposants")
    ws.append([HEAD_FR[c] for c in COLS])
    hfill = PatternFill("solid", fgColor="1F4E5F")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = hfill
        cell.alignment = Alignment(vertical="center")
    okfill = PatternFill("solid", fgColor="E5F5E0")
    chkfill = PatternFill("solid", fgColor="FFF3CD")
    for r in rows:
        ws.append([r[c] for c in COLS])
        st = r["statut_validation"]
        if st == "valide":
            ws.cell(ws.max_row, 6).fill = okfill
        elif st in ("a_verifier", "a_enrichir"):
            ws.cell(ws.max_row, 6).fill = chkfill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{ws.max_row}"
    widths = {"entreprise": 30, "secteur_metier": 28, "site_web": 30, "email": 30,
              "email_site_source": 30, "source_url": 32, "salon": 12}
    for i, c in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 14)

    wb.save("base-outbound.xlsx")
    print(f"WROTE base-outbound.xlsx | {len(rows)} exposants | {n_mail} emails | "
          f"{len(salons)} salon(s)")


if __name__ == "__main__":
    main()
